# -*- coding: utf-8 -*-
#!/usr/bin/env python3

from __future__ import absolute_import, division, print_function

from copy import copy

import os
import re
import sys

# Read and process MS Excel documents from on script into another
# Unicode characters.

# TODO: define data ranges (worksheet and cell ranges) on command line,
# then apply changes to only those cells.
# TODO: Create and add named style for the newly included font.

# https://openpyxl.readthedocs.io/en/default/tutorial.html
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font

import convertUtil

import adlamConversion

class cellData():
    def __init__(self):
        self.text = ''

class convertWorkbook():
    def __init__(self, input_path, output_dir, converter, debug=False, cell_ranges=None):
        self.oldFonts = []
        self.input_path = input_path
        self.output_dir = output_dir
        self.converter = converter
        self.old_fonts = converter.oldFonts  # List of font names
        self.unicode_font = converter.unicodeFont

        self.cell_ranges = cell_ranges

        self.workbook = None

        if self.input_path:
            self.workbook = load_workbook(filename=self.input_path)
        # The sheets
        for sheet in self.workbook:
            print(sheet.title)

    def process(self):
      
        print('process path = %s, output_dir = %s\n' % (
          self.input_path, self.output_dir))

        # Get the first sheet
        sheets = self.workbook.sheetnames
        ws = self.workbook[sheets[0]]

        # Get the output font and create a new named style for it
        new_style = NamedStyle('ConvertedFont')
        new_font = Font()
        new_font.name = self.converter.defaultOutputFont
        new_style.font = new_font
        self.workbook.add_named_style(new_style)

        self.converter.setScriptIndex(adlamConversion.LATIN2ADLAM)
        # Get the range to convert.
        # for cell in self.cells_to_convert:
        #     latin = ws[cell].value
        #     adlam = self.converter.convertText(latin, fontIndex=adlamConversion.LATIN2ADLAM)
        #     ws[cell] = adlam
        # Special case
        cells = []
        for range1 in self.cell_ranges:
            this_range = ws[range1[0]:range1[1]]
            cells.extend(this_range)

        for cell in cells:
            the_cell = cell[0]
            latin = the_cell.value
            if not latin:
                # Nothing there to process
                continue
            adlam = self.converter.convertText(latin, fontIndex=adlamConversion.LATIN2ADLAM)
            the_cell.value = adlam
            # TODO: Set the font
            old_font = the_cell.font
            new_font = copy(old_font)
            new_font.name = self.converter.defaultOutputFont
            the_cell.style = new_style
            the_cell.font = new_font

    def processText(self):
        # Get the converter
        # For each item in the range specified, check the font information

        # Convert contents if needed

        # For each sheet, create copy
        # For each cell, convert the cell(s) as neeed

        # Save the new .xlsx file.
        return


    def info(self):
        # Get the workbook
        # chartsheets: list of all of them
        self.workbook = Woorbook
        #
        return


def parse_cell_ranges(cell_string):
    # Get list of cell ranges from comma separated list of cells and ranges
    # also, ignore white space
    split_ranges_pattern = re.compile(r'[\s\,\;]+')
    rough_ranges = split_ranges_pattern.split(cell_string)
    ranges = []
    for g in rough_ranges:
        if g.find(':') >= 0:
            gs = g.split(':')
            ranges.append(gs)
        else:
            ranges.append([g, g])
    return ranges

# For standalone and testing.
def main(argv):
    global debug_output

    args = convertUtil.parseArgs()

    # Ranges of cells, e.g., "c2:c100, d2, e3:g9]"
    # Separated by comma, semicolon, or space
    cell_ranges = parse_cell_ranges(args.cells)

    debug_output = True
    paths_to_doc = args.filenames
    print('ARGS = %s' % args)

    for path in paths_to_doc:
        extension = os.path.splitext(path)[-1]

        converter = adlamConversion.AdlamConverter()
        if extension == '.xlsx':
            out_file_name = converter.get_outfile_name(path)  # Temporary
            processor = convertWorkbook(path, out_file_name, converter, debug_output, cell_ranges=cell_ranges)

            processor.process()  # Do the requested conversion

            processor.workbook.save(out_file_name)
        else:
            print('!!! Not processing file %s !' % path)


if __name__ == "__main__":
    main(sys.argv)
